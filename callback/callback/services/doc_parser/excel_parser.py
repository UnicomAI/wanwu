"""Excel 解析器：xlsx(openpyxl) 与 xls(xlrd)，输出 markdown 表格。

输出规则：
- 多行：第一行作表头，其余作数据行，渲染为标准 markdown 表格；
- 单行单列：无表头可用，直接输出原始值；
- 单行多列：该行作表头行，渲染为只有表头的 markdown 表格；
- 多 sheet：每个有数据的 sheet 渲染成一个独立表格，表格前加 ``### {sheet名}``
  标题区分；单 sheet 直接输出表格，不加标题。

单元格内的竖线转义、换行替换为空格，避免破坏表格结构；
表头列与数据行统一对齐到第一行列数，缺列补空，尾部全空列裁掉。

xlsx 处理合并单元格：合并区域内非左上角的空单元格用左上角值填充
（仅纵向合并向下填充，横向合并不重复）。xls(xlrd) 不处理合并单元格。
"""

from openpyxl import load_workbook
import xlrd

from .base import DocumentParser


def _escape_cell(value) -> str:
    """单元格值转 markdown 表格单元文本：None→空，整值浮点去尾零，竖线转义，换行→空格。"""
    if value is None:
        return ""
    # xlrd/openpyxl 把整数存成 float（如 20 → 20.0），整值浮点转回整数文本
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).replace("\\", "\\\\").replace("|", "\\|").replace("\n", " ").strip()


def _trim_trailing_empty(matrix: list) -> list:
    """裁掉所有行尾部均为空的列，避免 markdown 表格出现无意义空列。"""
    if not matrix:
        return matrix
    width = max(len(r) for r in matrix)
    last_non_empty = -1
    for c in range(width):
        if any(c < len(r) and r[c] not in (None, "") for r in matrix):
            last_non_empty = c
    keep = last_non_empty + 1
    if keep <= 0:
        return [[] for _ in matrix]
    return [r[:keep] for r in matrix]


def _rows_to_markdown(rows: list, table_headers: list) -> list:
    """把已转义的行列表拼成 markdown 表格文本块。

    Args:
        rows: 已 escape 的单元格字符串行列表（含表头行）。
        table_headers: 表头列名（已 escape），长度决定表格列数。

    Returns:
        拼好的 markdown 表格行列表（每行带 ``\\n``）。
    """
    ncol = len(table_headers)
    lines = []
    lines.append("| " + " | ".join(table_headers) + " |\n")
    lines.append("| " + " | ".join(["---"] * ncol) + " |\n")
    for r in rows:
        # 对齐到表头列数：不足补空，多余截断
        cells = (r + [""] * ncol)[:ncol]
        lines.append("| " + " | ".join(cells) + " |\n")
    return lines


def _build_merged_value_map(ws) -> dict:
    """构建 {(row, col): 值} 映射，仅对纵向合并向下填充。

    只有纵向合并(同一列跨多行)会被向下填充；横向合并(同一行跨多列)
    的右侧空单元格保持为空，不重复填充。
    """
    merged_map = {}
    for rng in ws.merged_cells.ranges:
        # 仅处理纵向合并：跨多行才向下填充；横向合并(max_row==min_row)不填充
        if rng.max_row <= rng.min_row:
            continue
        top_left_value = ws.cell(rng.min_row, rng.min_col).value
        if top_left_value is None:
            continue
        # 向下填充：合并区域内 min_row 之下的各行(同列)
        for r in range(rng.min_row + 1, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged_map[(r, c)] = top_left_value
    return merged_map


class XlsxParser(DocumentParser):
    def parse(self, file_path: str, max_tokens: int = 0) -> str:
        limit = self.max_chars(max_tokens)
        parts = []
        total = 0

        def append(text: str) -> bool:
            """追加文本；达到上限返回 True 表示应提前终止。"""
            nonlocal total
            parts.append(text)
            total += len(text)
            return limit is not None and total >= limit

        # data_only=True 取公式计算后的值；不用 read_only 以支持 merged_cells.ranges
        wb = load_workbook(file_path, data_only=True)
        try:
            # 先收集每个 sheet 的非空行；multi 由“有数据的 sheet 数”决定
            sheets_with_rows = []
            for ws in wb.worksheets:
                rows = [
                    row for row in ws.iter_rows()
                    if any(c.value is not None for c in row)
                ]
                if rows:
                    sheets_with_rows.append((ws, rows))
            multi = len(sheets_with_rows) > 1

            for ws, rows in sheets_with_rows:
                # 收集原始值矩阵：合并单元格空位用左上角值填充
                merged_map = _build_merged_value_map(ws)
                matrix = []
                for row in rows:
                    vals = []
                    for cell in row:
                        val = cell.value
                        if val is None:
                            val = merged_map.get((cell.row, cell.column))
                        vals.append(val)
                    matrix.append(vals)

                if append(self._render_sheet(ws.title, matrix, multi)):
                    return self.truncate("".join(parts), max_tokens)
        finally:
            wb.close()
        return self.truncate("".join(parts), max_tokens)

    @staticmethod
    def _render_sheet(sheet_name: str, matrix: list, multi: bool) -> str:
        """把单个 sheet 的原始值矩阵渲染为 markdown 文本块。"""
        matrix = _trim_trailing_empty(matrix)
        ncol = max((len(r) for r in matrix), default=0)

        # 单行单列：直接输出原始值，不套表格
        if len(matrix) == 1 and ncol <= 1:
            return _escape_cell(matrix[0][0] if matrix[0] else None) + "\n"

        # 表头取第一行；数据行取其余
        headers = [_escape_cell(v) for v in matrix[0]] if matrix else []
        data_rows = [[_escape_cell(v) for v in r] for r in matrix[1:]]
        lines = []
        if multi:
            lines.append(f"### {sheet_name}\n\n")
        lines.extend(_rows_to_markdown(data_rows, headers))
        lines.append("\n")
        return "".join(lines)


class XlsParser(DocumentParser):
    def parse(self, file_path: str, max_tokens: int = 0) -> str:
        limit = self.max_chars(max_tokens)
        parts = []
        total = 0

        def append(text: str) -> bool:
            nonlocal total
            parts.append(text)
            total += len(text)
            return limit is not None and total >= limit

        wb = xlrd.open_workbook(file_path)
        sheets = wb.sheets()
        valid_sheets = [
            s for s in sheets
            if any(any(c.value not in ("", None) for c in s.row(i)) for i in range(s.nrows))
        ]
        multi = len(valid_sheets) > 1
        for sheet in valid_sheets:
            rows = [
                [c.value for c in sheet.row(i)]
                for i in range(sheet.nrows)
            ]
            rows = [r for r in rows if any(c not in ("", None) for c in r)]
            if not rows:
                continue
            if append(XlsxParser._render_sheet(sheet.name, rows, multi)):
                return self.truncate("".join(parts), max_tokens)
        return self.truncate("".join(parts), max_tokens)
