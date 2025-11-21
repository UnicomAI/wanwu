import json
import openpyxl
import pandas as pd
import requests
answer_list = []
# Initialize一个Empty字符串来存储组合After的output  
answer_output = ""  
column_name=['RAG-searchList','RAG-output']


excel_file_path_label =r"RAG-test.xlsx"
excel_file_path_save = r"RAG-test-result.xlsx"

# OpenExcelFile
workbook = openpyxl.load_workbook(excel_file_path_label)
# 选择指定的工作Table
sheet = workbook['RAG基线数据集']   

# 读取ExcelFileData
df = pd.read_excel(excel_file_path_label,usecols = ['序号','问题'])
# 提取所需的两列Data
question_list = df['问题'].astype(str).tolist()

# 设置URL  
url = 'http://localhost:10891/rag/knowledge/stream/search'
user_id = "1"
kb_name = ['rag_base_test']
headers = {
            "Content-Type": "application/json",
            "X-uid": user_id
            }


for i,data in enumerate(column_name):
    column_index = 1
    while sheet.cell(row=1, column=column_index).value is not None:
                column_index += 1
    # 写入列标题
    sheet.cell(row=1, column=column_index, value=data)
     #Knowledge Base名设置为切分Length，分别为125/250/500
for j,question in enumerate(question_list):
    payload = json.dumps({
            "knowledgeBase":kb_name, 
            "question":question,
            "threshold":0,
            "topK":5,
            "history":[],
            "stream":True,
            "search_field":"emc",
            })
    # print(payload)
    response = requests.request("POST", url, headers=headers, data=payload,verify=False)
    
    
    
    answer_output = ""  
    answer_searchList = None  
  
    for line in response.iter_lines():  
        if line.startswith(b"data: "):  
            # 去除Before缀并解码为str  
            json_str = line[6:].decode('utf-8')  
            try:  
                # ParseJSON字符串为Python字典  
                data = json.loads(json_str)  
                # 提取outputField的值，并添加到answer_outputMedium  
                answer_output += data['data']['output']  
                # 提取searchList，If它IsNon-empty的且answer_searchList尚未设置  
                if data['data']['searchList'] and not answer_searchList:  
                    answer_searchList = data['data']['searchList']  
            except (json.JSONDecodeError, KeyError):  
                # IfParseFailedOr缺Few关键Field，则忽略OrRecordError  
                pass  # Or打印ErrorInfo  
  
    # Output组合After的output字符串  
    print(answer_output)  
  
    # If找到了Non-empty的searchList，则打印它（在这个例ChildMedium它将IsNone）  
    if answer_searchList:  
        print(answer_searchList)  
    else:  
        print("没有找到非空的searchList")
    
    answer_list.clear()
    answer_list.append(answer_searchList)  # 将Result添加到ListMedium  
    answer_list.append(answer_output)  # 将Result添加到ListMedium  
    for i,data in enumerate(answer_list):
        # 确定New列的Index:标题行Left起第一个为Empty的列
        column_index = 1
        while sheet.cell(row=j+2, column=column_index).value is not None:
            column_index += 1
        # 写入Data`
        json_data = json.dumps(data, ensure_ascii=False)  
        sheet.cell(row=j+2, column=column_index, value=json_data)  
   
        # SaveModifyAfter的ExcelFile
        workbook.save(excel_file_path_save)
        workbook.close()     
print("全部测试完成，请查看结果")
